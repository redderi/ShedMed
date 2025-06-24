from openpyxl import load_workbook
import pdfplumber
from sqlalchemy import inspect
from PySide6.QtWidgets import QMessageBox
import logging
from bd.equipment_model import Equipment
from utils.date_utils import parse_date

logger = logging.getLogger(__name__)

def import_from_excel(window, file_path, replace_table):
    logger.debug(f"Импорт данных из Excel: {file_path}, режим: {'Замена' if replace_table else 'Добавление'}")
    try:
        wb = load_workbook(file_path)
        expected_columns = 11
        fields = [
            "id", "equipment_name", "equipment_type", "serial_number", "location",
            "calibration_interval", "last_calibration_date", "next_calibration_date",
            "certificate_number", "days_to_calibration", "notes"
        ]
        
        errors = []
        table_columns = {col["name"] for col in inspect(window.session_manager.bind).get_columns('equipment')}
        row_count = 0
        
        window.ui.searchLineEdit.textChanged.disconnect()
        window.ui.searchColumnComboBox.currentTextChanged.disconnect()
        
        try:
            if replace_table:
                logger.debug("Очистка таблицы equipment")
                window.session_manager.query(Equipment).delete()
                window.session_manager.commit()
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                logger.debug(f"Обработка листа: {sheet_name}")
                header_found = False
                
                for row_idx, row in enumerate(ws.iter_rows(min_row=1), start=1):
                    row_values = [cell.value for cell in row]
                    if len(row_values) != expected_columns:
                        logger.warning(f"Лист {sheet_name}, строка {row_idx}: некорректное количество колонок ({len(row_values)} вместо {expected_columns})")
                        errors.append(f"Лист {sheet_name}, строка {row_idx}: ожидается {expected_columns} колонок, найдено {len(row_values)}")
                        continue
                    
                    logger.debug(f"Лист {sheet_name}, обработка строки {row_idx}: {row_values}")
                    
                    is_header = all(isinstance(v, str) for v in row_values if v is not None)
                    if is_header and not header_found:
                        logger.debug(f"Лист {sheet_name}, строка {row_idx} определена как заголовки: {row_values}")
                        header_found = True
                        continue
                    
                    if header_found:
                        data = {}
                        try:
                            for i, (field, value) in enumerate(zip(fields, row_values)):
                                if field == "id":
                                    continue
                                
                                value = str(value).strip() if value is not None else None
                                
                                if field in ["equipment_name", "serial_number"]:
                                    if not value:
                                        raise ValueError(f"Поле '{field}' не может быть пустым")
                                    data[field] = value
                                
                                elif field == "calibration_interval":
                                    if not value:
                                        raise ValueError(f"Поле 'calibration_interval' не может быть пустым")
                                    try:
                                        data[field] = int(value)
                                    except ValueError:
                                        raise ValueError(f"Неверный формат для 'calibration_interval': {value}")
                                
                                elif field == "last_calibration_date":
                                    if not value:
                                        raise ValueError(f"Поле 'last_calibration_date' не может быть пустым")
                                    try:
                                        data[field] = parse_date(value)
                                    except ValueError as e:
                                        raise ValueError(f"Неверный формат даты для 'last_calibration_date': {str(e)}")
                                
                                elif field == "next_calibration_date":
                                    if value:
                                        try:
                                            data[field] = parse_date(value)
                                        except ValueError as e:
                                            raise ValueError(f"Неверный формат даты для 'next_calibration_date': {str(e)}")
                                    else:
                                        data[field] = None
                                
                                else:
                                    data[field] = value
                            
                            required_fields = ["equipment_name", "serial_number", "calibration_interval", "last_calibration_date"]
                            missing_fields = [f for f in required_fields if not data.get(f)]
                            if missing_fields:
                                raise ValueError(f"Отсутствуют обязательные поля: {', '.join(missing_fields)}")
                            
                            new_equipment = Equipment()
                            try:
                                from handlers.equipment_handlers import process_equipment_data
                                process_equipment_data(window, new_equipment, data, table_columns)
                                new_equipment.update_calibration()
                                window.session_manager.add(new_equipment)
                                row_count += 1
                                logger.debug(f"Лист {sheet_name}, строка {row_idx} успешно обработана")
                            except Exception as e:
                                raise ValueError(f"Ошибка обработки данных: {str(e)}")
                        
                        except ValueError as e:
                            errors.append(f"Лист {sheet_name}, строка {row_idx}: {str(e)}")
                            logger.warning(f"Ошибка в листе {sheet_name}, строке {row_idx}: {str(e)}")
                            continue
                
            if errors:
                window.session_manager.rollback()
                error_msg = "Импорт завершён с ошибками:\n" + "\n".join(errors[:5])
                if len(errors) > 5:
                    error_msg += f"\n...и ещё {len(errors) - 5} ошибок"
                logger.error(error_msg)
                window.show_error("Ошибки импорта", error_msg)
            else:
                window.session_manager.commit()
                logger.info(f"Успешно импортировано {row_count} записей")
                QMessageBox.information(window, "Успех", f"Импортировано {row_count} записей")
                window.last_selected_row = None
                window.last_selected_event_row = None
                window.model.refresh()
                window.ui.tableView.resizeRowsToContents()
                window.apply_sorting()
        
        finally:
            window.ui.searchLineEdit.textChanged.connect(window.on_search_text_changed)
            window.ui.searchColumnComboBox.currentTextChanged.connect(window.on_search_text_changed)
    
    except Exception as e:
        window.session_manager.rollback()
        logger.error(f"Ошибка при импорте из Excel: {str(e)}")
        window.show_error("Ошибка", f"Ошибка при импорте из Excel: {str(e)}")

def import_from_pdf(window, file_path, replace_table):
    logger.debug(f"Импорт данных из PDF: {file_path}, режим: {'Замена' if replace_table else 'Добавление'}")
    try:
        with pdfplumber.open(file_path) as pdf:
            if not pdf.pages:
                logger.error("PDF не содержит страниц")
                window.show_error("Ошибка", "PDF не содержит страниц")
                return
            
            expected_columns = 11
            fields = [
                "id", "equipment_name", "equipment_type", "serial_number", "location",
                "calibration_interval", "last_calibration_date", "next_calibration_date",
                "certificate_number", "days_to_calibration", "notes"
            ]
            
            errors = []
            table_columns = {col["name"] for col in inspect(window.session_manager.bind).get_columns('equipment')}
            row_count = 0
            
            window.ui.searchLineEdit.textChanged.disconnect()
            window.ui.searchColumnComboBox.currentTextChanged.disconnect()
            
            try:
                if replace_table:
                    logger.debug("Очистка таблицы equipment")
                    window.session_manager.query(Equipment).delete()
                    window.session_manager.commit()
                
                for page_idx, page in enumerate(pdf.pages, start=1):
                    tables = page.extract_tables()
                    if not tables:
                        logger.debug(f"Страница {page_idx}: таблицы не найдены")
                        continue
                    
                    for table_idx, table in enumerate(tables, start=1):
                        logger.debug(f"Обработка таблицы {table_idx} на странице {page_idx}")
                        header_found = False
                        
                        for row_idx, row in enumerate(table, start=1):
                            if len(row) != expected_columns:
                                logger.warning(f"Страница {page_idx}, таблица {table_idx}, строка {row_idx}: некорректное количество колонок ({len(row)} вместо {expected_columns})")
                                errors.append(f"Страница {page_idx}, таблица {table_idx}, строка {row_idx}: ожидается {expected_columns} колонок, найдено {len(row)}")
                                continue
                            
                            logger.debug(f"Страница {page_idx}, таблица {table_idx}, обработка строки {row_idx}: {row}")
                            
                            is_header = all(isinstance(v, str) for v in row if v is not None)
                            if is_header and not header_found:
                                logger.debug(f"Страница {page_idx}, таблица {table_idx}, строка {row_idx} определена как заголовки: {row}")
                                header_found = True
                                continue
                            
                            if header_found:
                                data = {}
                                try:
                                    for i, (field, value) in enumerate(zip(fields, row)):
                                        if field == "id":
                                            continue
                                        
                                        value = str(value).replace("\n", " ").strip() if value else None
                                        
                                        if field in ["equipment_name", "serial_number"]:
                                            if not value:
                                                raise ValueError(f"Поле '{field}' не может быть пустым")
                                            data[field] = value
                                        
                                        elif field == "calibration_interval":
                                            if not value:
                                                raise ValueError(f"Поле 'calibration_interval' не может быть пустым")
                                            try:
                                                data[field] = int(value)
                                            except ValueError:
                                                raise ValueError(f"Неверный формат для 'calibration_interval': {value}")
                                        
                                        elif field == "last_calibration_date":
                                            if not value:
                                                raise ValueError(f"Поле 'last_calibration_date' не может быть пустым")
                                            try:
                                                data[field] = parse_date(value)
                                            except ValueError as e:
                                                raise ValueError(f"Неверный формат даты для 'last_calibration_date': {str(e)}")
                                        
                                        elif field == "next_calibration_date":
                                            if value:
                                                try:
                                                    data[field] = parse_date(value)
                                                except ValueError as e:
                                                    raise ValueError(f"Неверный формат даты для 'next_calibration_date': {str(e)}")
                                            else:
                                                data[field] = None
                                        
                                        else:
                                            data[field] = value
                                    
                                    required_fields = ["equipment_name", "serial_number", "calibration_interval", "last_calibration_date"]
                                    missing_fields = [f for f in required_fields if not data.get(f)]
                                    if missing_fields:
                                        raise ValueError(f"Отсутствуют обязательные поля: {', '.join(missing_fields)}")
                                    
                                    new_equipment = Equipment()
                                    try:
                                        from handlers.equipment_handlers import process_equipment_data
                                        process_equipment_data(window, new_equipment, data, table_columns)
                                        new_equipment.update_calibration()
                                        window.session_manager.add(new_equipment)
                                        row_count += 1
                                        logger.debug(f"Страница {page_idx}, таблица {table_idx}, строка {row_idx} успешно обработана")
                                    except Exception as e:
                                        raise ValueError(f"Ошибка обработки данных: {str(e)}")
                                
                                except ValueError as e:
                                    errors.append(f"Страница {page_idx}, таблица {table_idx}, строка {row_idx}: {str(e)}")
                                    logger.warning(f"Ошибка на странице {page_idx}, таблица {table_idx}, строка {row_idx}: {str(e)}")
                                    continue
                
                if errors:
                    window.session_manager.rollback()
                    error_msg = "Импорт завершён с ошибками:\n" + "\n".join(errors[:5])
                    if len(errors) > 5:
                        error_msg += f"\n...и ещё {len(errors) - 5} ошибок"
                    logger.error(error_msg)
                    window.show_error("Ошибки импорта", error_msg)
                else:
                    window.session_manager.commit()
                    logger.info(f"Успешно импортировано {row_count} записей")
                    QMessageBox.information(window, "Успех", f"Импортировано {row_count} записей")
                    window.last_selected_row = None
                    window.last_selected_event_row = None
                    window.model.refresh()
                    window.ui.tableView.resizeRowsToContents()
                    window.apply_sorting()
            
            finally:
                window.ui.searchLineEdit.textChanged.connect(window.on_search_text_changed)
                window.ui.searchColumnComboBox.currentTextChanged.connect(window.on_search_text_changed)
        
    except Exception as e:
        window.session_manager.rollback()
        logger.error(f"Ошибка при импорте из PDF: {str(e)}")
        window.show_error("Ошибка", f"Ошибка при импорте из PDF: {str(e)}")
