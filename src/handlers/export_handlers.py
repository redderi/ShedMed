from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.styles import getSampleStyleSheet
from PySide6.QtWidgets import QFileDialog
from PySide6.QtCore import QUrl
from PySide6.QtGui import QDesktopServices
import os
from datetime import datetime
from PySide6.QtCore import Qt
import logging
from PySide6.QtWidgets import QMessageBox
from utils.text_utils import insert_soft_hyphens
from utils.path_utils import resource_path

logger = logging.getLogger(__name__)

def export_data(window, format_type):
    logger.debug(f"Экспорт данных в формате: {format_type}")
    try:
        today = datetime.now().strftime("%Y-%m-%d")
        default_filename = f"equipment_export_{today}.{format_type}"
        
        file_dialog = QFileDialog(window)
        file_dialog.setAcceptMode(QFileDialog.AcceptSave)
        file_dialog.setDefaultSuffix(format_type)
        file_dialog.selectFile(default_filename)
        
        if format_type == "xlsx":
            file_dialog.setNameFilter("Excel Files (*.xlsx)")
        elif format_type == "pdf":
            file_dialog.setNameFilter("PDF Files (*.pdf)")
        
        if file_dialog.exec():
            file_path = file_dialog.selectedFiles()[0]
            logger.debug(f"Выбран путь для экспорта: {file_path}")
            
            headers = window.model.headers
            data = []
            for row in range(window.proxy_model.rowCount()):
                row_data = []
                for col in range(window.proxy_model.columnCount()):
                    index = window.proxy_model.index(row, col)
                    value = window.proxy_model.data(index, Qt.DisplayRole) or ""
                    row_data.append(value)
                data.append(row_data)
            
            if format_type == "xlsx":
                export_to_excel(window, file_path, headers, data)
            elif format_type == "pdf":
                export_to_pdf(window, file_path, headers, data)
            
            QMessageBox.information(window, "Успех", f"Данные успешно экспортированы в {file_path}")
            logger.info(f"Данные экспортированы в {file_path}")
            
            if format_type == "pdf" and os.path.exists(file_path):
                logger.debug(f"Открытие PDF для предпросмотра: {file_path}")
                QDesktopServices.openUrl(QUrl.fromLocalFile(file_path))
            
        else:
            logger.debug("Выбор файла для экспорта отменён")
    except Exception as e:
        logger.error(f"Ошибка при экспорте данных: {str(e)}")
        window.show_error("Ошибка", f"Ошибка при экспорте данных: {str(e)}")

def export_to_excel(window, file_path, headers, data):
    wb = Workbook()
    ws = wb.active
    ws.title = "Equipment Data"
    
    ws.append(headers)
    
    for row_data in data:
        ws.append(row_data)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min((max_length + 2) * 1.5, 30)
        ws.column_dimensions[column].width = adjusted_width
    
    wb.save(file_path)

def export_to_pdf(window, file_path, headers, data):
    try:
        font_path = resource_path("res/fonts/DejaVuSans.ttf")
        if not os.path.exists(font_path):
            logger.error(f"Шрифт {font_path} не найден")
            raise FileNotFoundError(f"Шрифт {font_path} не найден")
        pdfmetrics.registerFont(TTFont("Montserrat", font_path))

        doc = SimpleDocTemplate(file_path, pagesize=landscape(A4), leftMargin=8, rightMargin=8, topMargin=8, bottomMargin=8)
        elements = []

        styles = getSampleStyleSheet()
        header_style = styles["Normal"]
        header_style.fontName = "Montserrat"
        header_style.fontSize = 6
        header_style.alignment = 1
        header_style.leading = 7

        cell_style = styles["Normal"]
        cell_style.fontName = "Montserrat"
        cell_style.fontSize = 6
        cell_style.alignment = 1
        cell_style.leading = 7

        formatted_headers = [Paragraph(insert_soft_hyphens(str(h)), header_style) for h in headers]
        
        formatted_data = []
        for row_data in data:
            row = [Paragraph(insert_soft_hyphens(str(value)), cell_style) for value in row_data]
            formatted_data.append(row)

        table_data = [formatted_headers] + formatted_data

        col_widths = []
        default_widths = [30, 100, 60, 60, 90, 60, 60, 60, 150, 60, 90]
        for i in range(window.proxy_model.columnCount()):
            width = window.ui.tableView.columnWidth(i)
            if width <= 0:
                width = default_widths[i]
            col_widths.append(width)
        logger.debug(f"Ширины колонок из tableView (px): {col_widths}")

        total_available_width = landscape(A4)[0] - 16
        total_pixel_width = sum(col_widths)
        if total_pixel_width > 0:
            scale_factor = total_available_width / total_pixel_width
            col_widths = [w * scale_factor for w in col_widths]
        else:
            col_widths = [total_available_width / len(headers)] * len(headers)

        col_widths[0] = max(col_widths[0], 20)
        col_widths[8] = max(col_widths[8], 80)
        logger.debug(f"Ширины колонок в PDF (pt): {col_widths}")

        table = Table(table_data, colWidths=col_widths, repeatRows=1)

        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('FONTNAME', (0, 0), (-1, -1), 'Montserrat'),
            ('FONTSIZE', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 2),
            ('RIGHTPADDING', (0, 0), (-1, -1), 2),
            ('TOPPADDING', (0, 0), (-1, -1), 2),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))

        elements.append(table)
        doc.build(elements)
    except Exception as e:
        logger.error(f"Ошибка при экспорте в PDF: {str(e)}")
        window.show_error("Ошибка", f"Ошибка при экспорте в PDF: {str(e)}")
        raise